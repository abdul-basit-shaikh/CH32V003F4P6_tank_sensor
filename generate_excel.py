import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

def apply_styling(ws, title, header_color="1B365D", link_col_idx=8):
    # Styles
    font_title = Font(name="Arial", size=13, bold=True, color="FFFFFF")
    fill_title = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    
    font_header = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    fill_header = PatternFill(start_color="2C4D75", end_color="2C4D75", fill_type="solid")
    
    font_data = Font(name="Arial", size=9)
    font_data_bold = Font(name="Arial", size=9, bold=True)
    font_link = Font(name="Arial", size=9, color="0044CC", underline="single", bold=True)
    
    fill_even = PatternFill(start_color="F4F7FB", end_color="F4F7FB", fill_type="solid")
    fill_odd = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    
    thin_border_side = Side(border_style="thin", color="D0D7DE")
    border_data = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    thick_bottom_side = Side(border_style="medium", color="1B365D")
    border_header = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thick_bottom_side)
    
    max_col = ws.max_column
    max_row = ws.max_row
    
    # Title Row styling (Row 1)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    cell_title = ws.cell(row=1, column=1)
    cell_title.font = font_title
    cell_title.fill = fill_title
    cell_title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36
    
    # Header Row styling (Row 2)
    ws.row_dimensions[2].height = 28
    for col in range(1, max_col + 1):
        cell = ws.cell(row=2, column=col)
        cell.font = font_header
        cell.fill = fill_header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border_header
        
    # Data Rows styling (Row 3 to max_row)
    for row in range(3, max_row + 1):
        ws.row_dimensions[row].height = 24
        is_even = (row % 2 == 0)
        current_fill = fill_even if is_even else fill_odd
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = font_data
            cell.fill = current_fill
            cell.border = border_data
            
            # Alignments & Formats
            if col in [1, 6]:  # S.No, Qty
                cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col in [2]:   # Reference / Pin
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = font_data_bold
            elif col in [3]:   # Name
                cell.alignment = Alignment(horizontal="left", vertical="center")
                cell.font = font_data_bold
            elif col == link_col_idx: # Robu.in Link column
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = font_link
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
    # Auto-adjust column widths
    for col in range(1, max_col + 1):
        col_letter = get_column_letter(col)
        max_len = 0
        for row in range(2, max_row + 1):
            val = str(ws.cell(row=row, column=col).value or '')
            lines = val.split('\n')
            for line in lines:
                if len(line) > max_len:
                    max_len = len(line)
        # Set bounds for width
        col_width = max(max_len + 3, 11)
        if col == link_col_idx:
            col_width = 20
        elif col_width > 55:
            col_width = 55
        ws.column_dimensions[col_letter].width = col_width

# =========================================================================
# 1. CREATE TANK SENSOR EXCEL WORKBOOK
# =========================================================================
wb_tank = openpyxl.Workbook()

# Sheet 1: BOM List
ws_bom = wb_tank.active
ws_bom.title = "BOM_Components_List"

bom_headers = ["S.No", "Ref", "Component Name", "Value / Specification", "Package / Footprint", "Qty", "Purpose & Function (Kis Liye Lagta Hai)", "Robu.in Link", "Buying / Engineering Notes"]

bom_data = [
    [1, "U1", "Main Microcontroller (MCU)", "CH32V003F4P6 (RISC-V 32-bit, 48MHz, 16KB Flash, 2KB SRAM)", "TSSOP-20", 1, 
     "Poore tank sensor ka main brain. Probes read karta hai, deep sleep manage karta hai, battery calculate karta hai, aur LoRa se controller ko data bhejta hai.", 
     "https://robu.in/?s=CH32V003&post_type=product", 
     "Ultra-low power (~10µA sleep). WCH RISC-V core."],
     
    [2, "U2", "Wireless LoRa Module", "SX1278 Ra-02 (Ai-Thinker 433MHz)", "SMD-16 Module", 1, 
     "433MHz Long Range Radio. Tanki se controller tak 500m - 1km+ range me water level aur battery data wirelessly transmit karta hai.", 
     "https://robu.in/?s=SX1278+LoRa+Module+Ra-02&post_type=product", 
     "SPI Interface, 433MHz frequency band."],
     
    [3, "Q6", "P-Channel MOSFET", "AO3401A / SI2301 (30V, 4.2A, Low Rds(on))", "SOT-23", 1, 
     "Anti-Corrosion High-Side Switch. Paani ke common wire me 24/7 current nahi bhejta. Jab MCU reading leta hai sirf 10ms ke liye VCC pulse deta hai, baki waqt paani me 0V rehta hai taaki probes rust/electrolysis se kharab na hon.", 
     "https://robu.in/?s=AO3401A+MOSFET&post_type=product", 
     "Gate connected to PD3 via 100k pull-up."],
     
    [4, "Q1", "NPN Transistor (25% Probe)", "BC547 / 2N3904 / MMBT3904", "TO-92 / SOT-23", 1, 
     "25% Water Level Buffer. Paani 25% probe ko touch karta hai toh base trigger hokar Collector (PC0) ko LOW (0V) pull karta hai.", 
     "https://robu.in/?s=BC547+Transistor&post_type=product", 
     "Emitter to GND, Base via 100k resistor."],
     
    [5, "Q2", "NPN Transistor (50% Probe)", "BC547 / 2N3904 / MMBT3904", "TO-92 / SOT-23", 1, 
     "50% Water Level Buffer. 50% level touch hone par Collector (PC1) ko LOW pull karta hai.", 
     "https://robu.in/?s=BC547+Transistor&post_type=product", 
     "Emitter to GND, Base via 100k resistor."],
     
    [6, "Q3", "NPN Transistor (75% Probe)", "BC547 / 2N3904 / MMBT3904", "TO-92 / SOT-23", 1, 
     "75% Water Level Buffer. 75% level touch hone par Collector (PC2) ko LOW pull karta hai.", 
     "https://robu.in/?s=BC547+Transistor&post_type=product", 
     "Emitter to GND, Base via 100k resistor."],
     
    [7, "Q4", "NPN Transistor (100% Probe)", "BC547 / 2N3904 / MMBT3904", "TO-92 / SOT-23", 1, 
     "100% Water Level Buffer. 100% level touch hone par Collector (PC4) ko LOW pull karta hai.", 
     "https://robu.in/?s=BC547+Transistor&post_type=product", 
     "Emitter to GND, Base via 100k resistor."],
     
    [8, "Q5", "P-Channel MOSFET (Option B)", "AO3401A / SI2301", "SOT-23", "1 (Opt B)", 
     "Zero-Drop Reverse Battery Protection. 2x AA direct mode me battery ulti lagne par circuit ko instantly disconnect karta hai (Sirf 0.004V drop).", 
     "https://robu.in/?s=AO3401A+MOSFET&post_type=product", 
     "Gate to GND, Source to Bat+, Drain to Switch."],
     
    [9, "U3", "3.3V Low-Iq LDO (Option A)", "HT7333-2 (Holtek) / XC6206P332MR", "SOT-89 / SOT-23", "1 (Opt A)", 
     "Voltage Regulator. 4x AA (6V) battery setup me voltage ko solid 3.3V pe regulate karta hai (Quiescent current < 3µA).", 
     "https://robu.in/?s=HT7333&post_type=product", 
     "High efficiency LDO for 5-7 years battery life."],
     
    [10, "D2", "Step-down & Reverse Diode (Option A)", "1N4007 / M7 / 1N4002", "DO-41 / SMA", "1 (Opt A)", 
     "LDO Overvoltage Protection. 4x AA naye cells (6.4V) ko 0.7V drop karke 5.7V safe banata hai aur reverse polarity rokta hai.", 
     "https://robu.in/?s=1N4007+Diode&post_type=product", 
     "1A general purpose rectifier."],
     
    [11, "F1", "Resettable PTC Fuse", "500mA Resettable Fuse (Bourns / Littelfuse)", "1206 SMD", 1, 
     "Short Circuit Protection. Agar battery terminals ya board me short circuit ho toh battery ko blast ya drain hone se bachata hai.", 
     "https://robu.in/?s=1206+PTC+Fuse+500mA&post_type=product", 
     "Auto-reset when short is cleared."],
     
    [12, "C1", "Main Bulk Reservoir Capacitor", "100µF 16V / 25V (Electrolytic / Ceramic)", "Radial / 1206", 1, 
     "Energy Reservoir. Battery ke paas lagta hai. Jab LoRa module TX ke waqt sudden 120mA current spike leta hai, ye voltage drop nahi hone deta.", 
     "https://robu.in/?s=100uF+Capacitor&post_type=product", 
     "Essential for battery longevity & radio stability."],
     
    [13, "C2", "LoRa Bulk Capacitor", "10µF 10V/16V Ceramic", "0805 SMD", 1, 
     "LoRa Local Power Buffer. LoRa VCC/GND ke 3mm paas lagta hai for local current supply.", 
     "https://robu.in/?s=10uF+0805+SMD+Capacitor&post_type=product", 
     "Low ESR Ceramic capacitor."],
     
    [14, "C3", "LoRa RF Decoupling Capacitor", "100nF (104) 50V Ceramic", "0603 / 0805 SMD", 1, 
     "High-Frequency Noise Filter. LoRa module ke digital switching aur SPI clock noise ko ground karta hai.", 
     "https://robu.in/?s=100nF+0805+SMD+Capacitor&post_type=product", 
     "Place right at LoRa VCC/GND pins."],
     
    [15, "C4", "MCU Decoupling Capacitor", "100nF (104) 50V Ceramic", "0603 / 0805 SMD", 1, 
     "MCU Power Filter. CH32V003 MCU ke VDD (Pin 9) aur VSS (Pin 7) ke bilkul paas lagta hai.", 
     "https://robu.in/?s=100nF+0805+SMD+Capacitor&post_type=product", 
     "Clean digital supply for MCU."],
     
    [16, "C5", "MCU Reset Capacitor", "100nF (104) 50V Ceramic", "0603 / 0805 SMD", 1, 
     "Clean Power-On Reset. NRST pin ko power-on par clean reset pulse deta hai aur false resets rokta hai.", 
     "https://robu.in/?s=100nF+0805+SMD+Capacitor&post_type=product", 
     "Connected from NRST to GND."],
     
    [17, "C6-C9", "Probe Filter Capacitors", "4x 100nF (104) 50V Ceramic", "0603 / 0805 SMD", 4, 
     "Noise / ESD Filter. Tanki ke 25%, 50%, 75%, 100% lambe taaron se aane wale static, RF noise aur false triggering ko ground karte hain.", 
     "https://robu.in/?s=100nF+0805+SMD+Capacitor&post_type=product", 
     "Connected across each probe input to GND."],
     
    [18, "C10", "Common Line Filter Capacitor", "100nF (104) 50V Ceramic", "0603 / 0805 SMD", 1, 
     "Common Line Protection. Common sensor line ko ESD aur static spike se bachata hai.", 
     "https://robu.in/?s=100nF+0805+SMD+Capacitor&post_type=product", 
     "Connected from Common to GND."],
     
    [19, "R3", "NRST Pull-up Resistor", "10kΩ 5% (0.125W)", "0603 / 0805 SMD", 1, 
     "Reset Pin Stabilizer. CH32V003 ke NRST (Pin 4) ko VCC se pull-up rakhta hai taaki chip hang na ho.", 
     "https://robu.in/?s=10k+0805+SMD+Resistor&post_type=product", 
     "NRST to VDD."],
     
    [20, "R4", "LED Resistor", "1kΩ 5% (0.125W)", "0603 / 0805 SMD", 1, 
     "LED Current Limiter. Status LED (PD2) ko safe 1-2mA current deta hai taaki battery waste na ho.", 
     "https://robu.in/?s=1k+0805+SMD+Resistor&post_type=product", 
     "PD2 to LED Anode."],
     
    [21, "R5-R8", "Probe Base Resistors", "4x 100kΩ 5% (0.125W)", "0603 / 0805 SMD", 4, 
     "Transistor Base Current Limiter. Q1-Q4 BC547 ke base me safe micro-ampere current allow karte hain jab paani touch hota hai.", 
     "https://robu.in/?s=100k+0805+SMD+Resistor&post_type=product", 
     "Probe lines to Base pins."],
     
    [22, "R9", "MOSFET Gate Pull-up Resistor", "100kΩ 5% (0.125W)", "0603 / 0805 SMD", 1, 
     "Q6 Gate Lock. PD3 floating hone par ya deep sleep ke waqt Q6 Gate ko VCC se pull karke MOSFET ko 100% OFF rakhta hai.", 
     "https://robu.in/?s=100k+0805+SMD+Resistor&post_type=product", 
     "Q6 Gate to VCC."],
     
    [23, "D1", "Status Indicator LED", "3mm Green LED / 0805 SMD", "SMD / TH", 1, 
     "User Feedback. Pairing mode, reset indication, aur data transmission confirmation blink karne ke liye.", 
     "https://robu.in/?s=Green+LED+0805&post_type=product", 
     "Active HIGH from PD2."],
     
    [24, "SW1", "User Push Button", "6x6mm Tactile Switch", "SMD / TH", 1, 
     "Pairing & Reset. 3s hold = Reboot, 5s hold = Controller ke saath Pairing Mode.", 
     "https://robu.in/?s=Tactile+Push+Button+6x6&post_type=product", 
     "Active LOW connected to PD6."],
     
    [25, "SW2", "Power Slide Switch", "SPST Slide Switch 3-Pin", "Through-Hole", 1, 
     "Battery ON/OFF Switch. Board ko storage ya maintenance ke waqt physically switch off karne ke liye.", 
     "https://robu.in/?s=SPST+Slide+Switch&post_type=product", 
     "In series with battery positive line."],
     
    [26, "J1", "Battery Connector", "2-Pin JST-XH / 2.54mm Screw Terminal", "Through-Hole", 1, 
     "Battery pack (2x AA ya 4x AA holder) connect karne ke liye.", 
     "https://robu.in/?s=JST+XH+2+Pin+Connector&post_type=product", 
     "Polarized connector recommended."],
     
    [27, "J2", "SWD Programming Header", "1x3 Male Header (2.54mm)", "Through-Hole", 1, 
     "Firmware Flashing. WCH-LinkE programmer se code upload karne ke liye (VCC, SWIO/PD1, GND).", 
     "https://robu.in/?s=2.54mm+Male+Berg+Strip&post_type=product", 
     "Single-wire debug header."],
     
    [28, "J3", "Tank Sensor Connector", "1x5 Screw Terminal Block (3.81mm / 5.08mm)", "Through-Hole", 1, 
     "Tanki ke 5 taar connect karne ke liye: Common (VCC), 25%, 50%, 75%, 100%.", 
     "https://robu.in/?s=5+Pin+Screw+Terminal+Block&post_type=product", 
     "Heavy duty screw terminal."],
     
    [29, "J4", "Debug UART Header", "1x2 Male Header (2.54mm)", "Through-Hole", 1, 
     "Serial monitor debugging ke liye (PD5 TX + GND).", 
     "https://robu.in/?s=2.54mm+Male+Berg+Strip&post_type=product", 
     "Optional for debugging."],
     
    [30, "ANT1", "433MHz LoRa Antenna", "Spring Coil Antenna / IPEX SMA Whip", "433 MHz", 1, 
     "LoRa signal transmit karne ke liye.", 
     "https://robu.in/?s=433MHz+Spring+Antenna&post_type=product", 
     "Solder directly or use IPEX connector."],
     
    [31, "PROBES", "Water Level Probes", "SS304 / SS316 Stainless Steel Screws / Rods", "M3 / M4 Bolt", 5, 
     "Paani ke andar daalne ke liye rust-proof probes.", 
     "https://robu.in/?s=Stainless+Steel+Bolt+M3&post_type=product", 
     "Stainless steel (Do not use bare copper)."],
     
    [32, "PROG1", "WCH Programmer (Tool)", "WCH-LinkE Programmer / Debugger", "USB Tool", 1, 
     "CH32V003 MCU me code flash aur debug karne ke liye official hardware programmer.", 
     "https://robu.in/?s=WCH-LinkE+Programmer&post_type=product", 
     "Connects to J2 header (SWIO, VCC, GND)."]
]

ws_bom.cell(row=1, column=1, value="CH32V003F4P6 TANK WATER LEVEL SENSOR — COMPLETE BOM & ROBU.IN BUYING LIST")
ws_bom.append(bom_headers)

for r_idx, row in enumerate(bom_data, start=3):
    url = row[7]
    display_row = list(row)
    display_row[7] = "🔗 View on Robu.in"
    ws_bom.append(display_row)
    cell = ws_bom.cell(row=r_idx, column=8)
    cell.hyperlink = url

apply_styling(ws_bom, "CH32V003 Tank Sensor BOM & Robu Links", "0B3C5D", link_col_idx=8)

# Sheet 2: Pin Mapping
ws_pins = wb_tank.create_sheet(title="MCU_Pin_Mapping")
pin_headers = ["TSSOP-20 Pin", "Pin Name", "Connected Component / Signal", "Direction / Mode", "Description & Circuit Function"]
pin_data = [
    ["Pin 1", "PD4", "SX1278 LoRa RST", "Output", "LoRa Module Hardware Reset Line"],
    ["Pin 2", "PD5", "J4 Debug UART TX", "Output", "115200 Baud Serial Debug Print (Optional)"],
    ["Pin 3", "PD6", "SW1 Push Button", "Input (Pull-Up)", "Pairing / Reset / Reboot Button (Active LOW, EXTI Wakeup)"],
    ["Pin 4", "PD7", "NRST (R3 10k + C5 104)", "Input / Reset", "MCU Hardware Reset with RC Noise Filter"],
    ["Pin 5", "PA1", "NC (Internal Vref used)", "Analog Input", "Battery monitored via 1.20V Internal Bandgap (0µA drain)"],
    ["Pin 6", "PA2", "NC", "Spare GPIO", "Unconnected / Spare Test Pad"],
    ["Pin 7", "VSS", "Ground Plane (GND)", "Power", "Main System Ground"],
    ["Pin 8", "PD0", "NC", "Spare GPIO", "Unconnected / Spare Test Pad"],
    ["Pin 9", "VDD", "VCC Rail (3.0V - 3.3V) + C4 (104)", "Power", "Main MCU Power Input (with 100nF Decoupling Cap)"],
    ["Pin 10", "PC0", "Q1 Collector (25% Probe)", "Input (Pull-Up)", "25% Water Level Detector (Reads LOW when water reaches 25%)"],
    ["Pin 11", "PC1", "Q2 Collector (50% Probe)", "Input (Pull-Up)", "50% Water Level Detector (Reads LOW when water reaches 50%)"],
    ["Pin 12", "PC2", "Q3 Collector (75% Probe)", "Input (Pull-Up)", "75% Water Level Detector (Reads LOW when water reaches 75%)"],
    ["Pin 13", "PC3", "SX1278 LoRa NSS (CS)", "Output", "SPI Chip Select for SX1278 LoRa Module"],
    ["Pin 14", "PC4", "Q4 Collector (100% Probe)", "Input (Pull-Up)", "100% Water Level Detector (Reads LOW when water reaches 100%)"],
    ["Pin 15", "PC5", "SX1278 LoRa SCK", "AF Output", "SPI1 Clock Line for LoRa Communication"],
    ["Pin 16", "PC6", "SX1278 LoRa MOSI", "AF Output", "SPI1 Master-Out-Slave-In Data Line"],
    ["Pin 17", "PC7", "SX1278 LoRa MISO", "Input (Pull-Up)", "SPI1 Master-In-Slave-Out Data Line"],
    ["Pin 18", "PD1", "J2 SWIO Header", "Bidirectional", "Single-Wire Debug / Flash Programming Interface"],
    ["Pin 19", "PD2", "D1 Status LED (via R4 1k)", "Output", "Status Indicator LED (Active HIGH: Blinks for TX, Pairing)"],
    ["Pin 20", "PD3", "Q6 P-MOSFET Gate (via R9 100k)", "Output", "Common Probe Driver (Active LOW: Pulsed 3V for 10ms)"]
]
ws_pins.cell(row=1, column=1, value="CH32V003F4P6 TSSOP-20 COMPLETE PIN MAPPING")
ws_pins.append(pin_headers)
for row in pin_data:
    ws_pins.append(row)
apply_styling(ws_pins, "MCU Pin Mapping", "164E63", link_col_idx=-1)

# Save Tank Sensor Workbook safely
tank_paths = [
    r"c:\Users\shaik\mounriver-studio-projects\CH32V003F4P6_tank_sensor\Tank_Sensor_Hardware_List.xlsx",
    r"c:\Users\shaik\mounriver-studio-projects\CH32V003F4P6_tank_sensor\Tank_Sensor_Hardware_BOM.xlsx"
]
for p in tank_paths:
    try:
        wb_tank.save(p)
        print(f"Saved Tank Sensor Excel: {p}")
        break
    except PermissionError:
        print(f"File {p} is open in Excel, trying alternate filename...")

# =========================================================================
# 2. CREATE WATER LEVEL CONTROLLER EXCEL WORKBOOK
# =========================================================================
wb_ctrl = openpyxl.Workbook()

# Sheet 1: Controller BOM
ws_ctrl_bom = wb_ctrl.active
ws_ctrl_bom.title = "Controller_BOM_List"

ctrl_bom_headers = ["S.No", "Ref / Identifier", "Component / IC Name", "Specification / Part No.", "Package / Type", "Qty", "Purpose & Function (Kis Liye Lagta Hai)", "Robu.in Link", "Buying / Hardware Assembly Notes"]

ctrl_bom_data = [
    [1, "U1", "Main Controller MCU", "ESP32-WROOM-32 / DevKit V1 (30/38 Pin)", "Dual-Core 240MHz, 4MB Flash", 1, 
     "Poore system ka main CPU. LoRa packets receive karta hai, 3 tanks ka logic run karta hai, OLED animation dikhata hai, RTC time schedule monitor karta hai, aur relays/buzzer control karta hai.", 
     "https://robu.in/?s=ESP32+Development+Board+38+Pin&post_type=product", 
     "ESP-IDF v5.x compatible. Integrated Wi-Fi & Bluetooth."],
     
    [2, "U2", "Wireless LoRa Module", "SX1278 Ra-02 (Ai-Thinker 433MHz)", "SMD Module / Breakout", 1, 
     "Wireless Receiver. Tank sensor se 433MHz frequency par encrypted level data, battery percentage, aur heartbeat signals receive karta hai.", 
     "https://robu.in/?s=SX1278+LoRa+Module+Ra-02&post_type=product", 
     "SPI Interface (SCK:18, MOSI:19, MISO:4, NSS:23, RST:22, DIO0:21)."],
     
    [3, "DISP1", "OLED Display Module", "1.3\" (SH1106) ya 0.96\" (SSD1306) I2C OLED", "128x64 Pixel I2C Module", 1, 
     "Main User Interface. Tank levels (25%, 50%, 75%, 100%), water wave animations, pump status, schedule time, dry-run timer, aur settings menu show karta hai.", 
     "https://robu.in/?s=1.3+inch+OLED+Display+I2C&post_type=product", 
     "I2C Address: 0x3C / 0x78 (SCL: GPIO 7, SDA: GPIO 6)."],
     
    [4, "U3", "High Precision RTC Module", "DS3231 High Precision I2C RTC", "SOIC-16 Breakout Module", 1, 
     "Real-Time Clock & Scheduling. Power cut hone ke baad bhi exact time yaad rakhta hai, taaki pump schedule (subah/shaam ka auto-on time) bina internet ke accurate chale.", 
     "https://robu.in/?s=DS3231+RTC+Module&post_type=product", 
     "Includes onboard temperature-compensated crystal (TCXO)."],
     
    [5, "BAT1", "RTC Backup Battery", "CR2032 3V Lithium Coin Cell", "20mm Coin Cell", 1, 
     "DS3231 RTC chip ko power cut ke waqt time run karne ke liye backup power deta hai (5+ saal life).", 
     "https://robu.in/?s=CR2032+Coin+Battery&post_type=product", 
     "Standard 3V lithium cell on RTC holder."],
     
    [6, "RLY1-3", "3-Channel Relay Module", "3-Channel 5V/12V Relay Board (Optocoupler Isolated)", "Module with Songle Relays", 1, 
     "Pump Switching (Pump 1, 2, 3). ESP32 ke 3.3V low-power signals se heavy AC motors / starter panels ko ON/OFF switch karta hai.", 
     "https://robu.in/?s=4+Channel+5V+Relay+Module+Optocoupler&post_type=product", 
     "Opto-isolated with JD-VCC jumper for noise immunity."],
     
    [7, "BZ1", "Active Buzzer", "5V DC Active Buzzer (Continuous Beep)", "12mm DIP Through-Hole", 1, 
     "Audio Alerts & Alarms. Tank full (100%), dry-run fault, sensor offline warning, aur button press key-tones provide karta hai.", 
     "https://robu.in/?s=5V+Active+Buzzer&post_type=product", 
     "Loud 85dB+ buzzer for audible alerts."],
     
    [8, "Q1", "Buzzer Driver Transistor", "SS8050 / 2N2222 / BC547 NPN", "TO-92 / SOT-23", 1, 
     "Buzzer Switch. ESP32 GPIO 5 (3.3V, 12mA max) se 5V buzzer ko safe drive karne ke liye buffer transistor.", 
     "https://robu.in/?s=2N2222+NPN+Transistor&post_type=product", 
     "NPN buffer switch."],
     
    [9, "D1", "Buzzer Flyback Diode", "1N4148 / 1N4007 Diode", "DO-35 / DO-41", 1, 
     "Flyback Protection Diode. Buzzer coil off hone par aane wale reverse inductive EMF spike se transistor aur ESP32 ko protect karta hai.", 
     "https://robu.in/?s=1N4148+Diode&post_type=product", 
     "Connected in reverse across buzzer pins."],
     
    [10, "R1", "Buzzer Base Resistor", "1kΩ 5% (0.25W)", "0805 / Axial", 1, 
     "Base Current Limiter. GPIO 5 se NPN transistor ke base me safe ~2.5mA current flow ensure karta hai.", 
     "https://robu.in/?s=1k+0805+SMD+Resistor&post_type=product", 
     "Connected between GPIO 5 and Base."],
     
    [11, "SW1-4", "User Input Buttons (4x)", "TTP223 Touch Modules / 6x6mm Tactile Buttons", "Module / SMD / TH", 4, 
     "User Controls: Btn 1: Menu/Select (GPIO 3), Btn 2: Pump 1 (GPIO 0), Btn 3: Pump 2 (GPIO 1), Btn 4: Pump 3 (GPIO 2).", 
     "https://robu.in/?s=TTP223+Touch+Sensor&post_type=product", 
     "Active HIGH signals. TTP223 capacitive touch recommended for modern look."],
     
    [12, "R2-R5", "Button Pull-Down Resistors", "4x 10kΩ 5% (0.125W)", "0805 / Axial", 4, 
     "Button Line Stabilizers. Passive mechanical switches ke liye button lines ko 0V pe hold rakhta hai (Floating noise prevent).", 
     "https://robu.in/?s=10k+0805+SMD+Resistor&post_type=product", 
     "Not needed if using TTP223 touch sensors."],
     
    [13, "R6-R7", "I2C Bus Pull-Up Resistors", "2x 4.7kΩ 5% (0.125W)", "0805 / Axial", 2, 
     "I2C Bus Pull-ups. SDA (GPIO 6) aur SCL (GPIO 7) lines ko stable 3.3V pull-up deta hai.", 
     "https://robu.in/?s=4.7k+Resistor&post_type=product", 
     "Built-in on most OLED/RTC breakout boards."],
     
    [14, "PS1", "Main AC-DC Power Supply", "230V AC to 5V 2A SMPS Adapter / Hi-Link HLK-PM01 (5V 3W/5W)", "Power Module / Adapter", 1, 
     "Main System Power. 230V AC mains ko clean 5V DC supply me convert karta hai poore controller ke liye.", 
     "https://robu.in/?s=Hi-Link+HLK-PM01+5V&post_type=product", 
     "Regulated 5V DC output with isolation."],
     
    [15, "VR1", "3.3V Voltage Regulator", "AMS1117-3.3 / MP1584 DC-DC Buck Module", "SOT-223 / Module", 1, 
     "Dedicated 3.3V Rail. 5V rail ko rock-solid 3.3V me convert karta hai ESP32 aur LoRa Module ke liye (800mA - 1A capability).", 
     "https://robu.in/?s=AMS1117+3.3V+Module&post_type=product", 
     "Prevents ESP32 brownout during Wi-Fi / LoRa TX."],
     
    [16, "C1", "LoRa Bulk Capacitor", "10µF 16V Ceramic / Tantalum", "0805 SMD", 1, 
     "LoRa Power Buffer. LoRa Ra-02 VCC/GND ke paas lagta hai, TX/RX current transients ko smooth karta hai.", 
     "https://robu.in/?s=10uF+0805+SMD+Capacitor&post_type=product", 
     "Essential for reliable radio reception."],
     
    [17, "C2-C4", "Decoupling Capacitors", "3x 100nF (104) 50V Ceramic", "0603 / 0805 SMD", 3, 
     "High Frequency Noise Filters. LoRa, ESP32, aur I2C bus ke paas high frequency noise ground karne ke liye.", 
     "https://robu.in/?s=100nF+0805+SMD+Capacitor&post_type=product", 
     "Place close to IC power pins."],
     
    [18, "C5", "Power Filter Bulk Cap", "470µF / 1000µF 16V Electrolytic", "Radial Through-Hole", 1, 
     "5V Rail Bulk Capacitor. 5V power input line par heavy filter capacitor taaki relays switch hone par voltage drop na ho.", 
     "https://robu.in/?s=1000uF+16V+Capacitor&post_type=product", 
     "Low ESR 105°C rated electrolytic."],
     
    [19, "SNUB1-3", "RC Snubbers (3x)", "0.1µF 400V (104) Cap + 100Ω 2W Resistor", "High Voltage Cap + 2W Resistor", 3, 
     "Relay Contact Spark / EMI Suppressor. Pump motor start/stop hone par Relay contacts par spark/EMI khatam karta hai taaki ESP32 reboot na ho.", 
     "https://robu.in/?s=RC+Snubber+Module&post_type=product", 
     "Connected in parallel across each Relay NO/COM contact."],
     
    [20, "MOV1-3", "MOVs (Varistors 3x)", "14D471K / 10D471K (470V MOV)", "Radial Disc 10mm/14mm", 3, 
     "AC Surge & Lightning Protection. AC Relay outputs par high voltage surges aur inductive voltage spikes ko suppress karta hai.", 
     "https://robu.in/?s=14D471K+MOV&post_type=product", 
     "Connected across AC Phase and Neutral."],
     
    [21, "TB1-4", "Screw Terminal Blocks", "2-Pin / 3-Pin (5.08mm / 7.62mm Pitch)", "PCB Mount Heavy Terminals", 4, 
     "AC 230V Mains Input aur Pump 1, Pump 2, Pump 3 load wires connect karne ke liye.", 
     "https://robu.in/?s=5.08mm+Screw+Terminal+Block&post_type=product", 
     "Rated for 250V AC 10A-16A."],
     
    [22, "ANT1", "433MHz High-Gain Antenna", "433MHz Whip / Rubber Duck Antenna with SMA Connector", "SMA 50Ω Antenna", 1, 
     "High-Gain External Antenna. Ground floor se rooftop tank tak maximum RF signal penetration ke liye.", 
     "https://robu.in/?s=433MHz+SMA+Antenna&post_type=product", 
     "SMA Male connector for wall-mount chassis."]
]

ws_ctrl_bom.cell(row=1, column=1, value="SMART WATER LEVEL CONTROLLER (ESP32) — BOM & ROBU.IN BUYING LIST")
ws_ctrl_bom.append(ctrl_bom_headers)

for r_idx, row in enumerate(ctrl_bom_data, start=3):
    url = row[7]
    display_row = list(row)
    display_row[7] = "🔗 View on Robu.in"
    ws_ctrl_bom.append(display_row)
    cell = ws_ctrl_bom.cell(row=r_idx, column=8)
    cell.hyperlink = url

apply_styling(ws_ctrl_bom, "ESP32 Controller BOM & Robu Links", "1E3A8A", link_col_idx=8)

# Sheet 2: Controller Pin Mapping
ws_ctrl_pins = wb_ctrl.create_sheet(title="ESP32_Pin_Mapping")
ctrl_pin_headers = ["ESP32 GPIO Pin", "Software Macro Name", "Connected Hardware / Module", "Interface & Direction", "Description & Hardware Function"]
ctrl_pin_data = [
    ["GPIO 18", "LORA_SCK_PIN", "SX1278 LoRa SCK", "SPI Clock Output", "SPI Clock for LoRa Communication"],
    ["GPIO 19", "LORA_MOSI_PIN", "SX1278 LoRa MOSI", "SPI Master Out", "SPI Master Out to LoRa Module"],
    ["GPIO 4", "LORA_MISO_PIN", "SX1278 LoRa MISO", "SPI Master In", "SPI Master In from LoRa Module"],
    ["GPIO 23", "LORA_NSS_PIN", "SX1278 LoRa NSS (CS)", "SPI Chip Select (Output)", "Active LOW Chip Select for LoRa"],
    ["GPIO 22", "LORA_RST_PIN", "SX1278 LoRa RST", "Output", "Hardware Reset Pin for LoRa Module"],
    ["GPIO 21", "LORA_DIO0_PIN", "SX1278 LoRa DIO0", "Input (Interrupt)", "Packet Received / Transmit Done Interrupt"],
    ["GPIO 7", "I2C_MASTER_SCL_IO", "OLED Display & DS3231 RTC SCL", "I2C Clock Output", "Shared I2C Clock Line (400kHz Fast Mode)"],
    ["GPIO 6", "I2C_MASTER_SDA_IO", "OLED Display & DS3231 RTC SDA", "I2C Data (Bidirectional)", "Shared I2C Data Line"],
    ["GPIO 15", "PUMP_PIN_1", "Relay Channel 1 (Pump 1)", "Output", "Pump 1 Relay Control (Active HIGH / LOW)"],
    ["GPIO 14", "PUMP_PIN_2", "Relay Channel 2 (Pump 2)", "Output", "Pump 2 Relay Control (Active HIGH / LOW)"],
    ["GPIO 17", "PUMP_PIN_3", "Relay Channel 3 (Pump 3)", "Output", "Pump 3 Relay Control (Active HIGH / LOW)"],
    ["GPIO 5", "BUZZER_PIN", "Buzzer Driver NPN Base", "Output", "Audio Alarm & Key Beep Control (via NPN Transistor)"],
    ["GPIO 3", "BUTTON_PIN_1", "Menu / Navigation Button", "Input (Active-HIGH)", "Menu navigation, settings & long press actions"],
    ["GPIO 0", "BUTTON_PIN_PUMP_1", "Pump 1 Manual Toggle Button", "Input (Active-HIGH)", "Manual ON/OFF Toggle for Pump 1"],
    ["GPIO 1", "BUTTON_PIN_PUMP_2", "Pump 2 Manual Toggle Button", "Input (Active-HIGH)", "Manual ON/OFF Toggle for Pump 2"],
    ["GPIO 2", "BUTTON_PIN_PUMP_3", "Pump 3 Manual Toggle Button", "Input (Active-HIGH)", "Manual ON/OFF Toggle for Pump 3"]
]

ws_ctrl_pins.cell(row=1, column=1, value="SMART WATER LEVEL CONTROLLER (ESP32) — GPIO PIN MAPPING")
ws_ctrl_pins.append(ctrl_pin_headers)
for row in ctrl_pin_data:
    ws_ctrl_pins.append(row)
apply_styling(ws_ctrl_pins, "ESP32 Pin Mapping", "065F46", link_col_idx=-1)

# Save Controller Workbook safely
ctrl_paths = [
    r"c:\Users\shaik\workspace\water_level_controller\Water_Level_Controller_Hardware_List.xlsx",
    r"c:\Users\shaik\workspace\water_level_controller\Water_Level_Controller_Hardware_BOM.xlsx"
]
for p in ctrl_paths:
    try:
        wb_ctrl.save(p)
        print(f"Saved Controller Excel: {p}")
        break
    except PermissionError:
        print(f"File {p} is open in Excel, trying alternate filename...")
